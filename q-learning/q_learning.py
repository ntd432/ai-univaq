import numpy as np
import matplotlib.pyplot as plt
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill

GRID_SIZE = (9, 4)
START_STATE = (4, 3)
GOAL_STATE = (8, 4)  
EGG_BEATER = [(1, 3), (8, 3)]
GATES = [(4, 2), (9, 3)]
ACTIONS = ['up', 'down', 'left', 'right', 'use_gate']
WALLS = [((1, 3), (1, 4)), ((1, 3), (2, 3)), ((2, 1), (1, 1)), ((2, 2), (2, 3)), 
         ((2, 2), (2, 1)), ((3, 2), (3, 3)), ((3, 2), (3, 1)), ((8, 3), (8, 4)), ((8, 3), (7, 3)), 
         ((9, 3), (9, 4)), ((8, 2), (8, 1)), ((8, 2), (7, 2))]  # Define walls as needed
EPISODES = 1000
ALPHA = 0.1  
GAMMA = 0.9
EPSILON = 0.1  

q_values = {(state, tool_status, action): 0 for state in [(x, y) for x in range(1, GRID_SIZE[0] + 1)
                                                          for y in range(1, GRID_SIZE[1] + 1)]
            for tool_status in [0, 1]  # 0: Tool not collected, 1: Tool collected
            for action in ACTIONS}

def is_valid(state, prev_state=None):
    x, y = state
    if x < 1 or x > GRID_SIZE[0] or y < 1 or y > GRID_SIZE[1]:
        return False
    # consider the walls
    if state[0] == 5:
        return False
    if (state, prev_state) in WALLS or (prev_state, state) in WALLS:
        return False
    return True

def get_next_state(state, action):
    x, y = state
    
    if action == 'use_gate':
        if state == (4, 2):  # Gate 1
            return (9, 3)
        elif state == (9, 3):  # Gate 2
            return (4, 2)
        else:
            return state  # No gate to use, stay in the same state

    # Standard movement actions
    if action == 'up':
        next_state = (x, y  + 1)
    elif action == 'down':
        next_state = (x, y - 1)
    elif action == 'left':
        next_state = (x - 1, y)
    elif action == 'right':
        next_state = (x + 1, y)
    else:
        next_state = state  # No move

    if not is_valid(next_state, prev_state=state):
        next_state = state  # Stay in the same position if invalid move

    return next_state

def manhattan_distance(state, goal):
    """Calculate Manhattan distance between two points."""
    return abs(state[0] - goal[0]) + abs(state[1] - goal[1])

def get_reward(state, next_state, tool_collected, action):
    if state == next_state:
        return -100
    if action == 'use_gate':
        if state == (4, 2):  # Positive reward for using gate at (4, 2)
            return 10
        elif state == (9, 3):  # Penalty for using gate at (9, 3)
            return -10
        else:
            return -100
    elif next_state == GOAL_STATE and tool_collected == 1:
        return 100  
    elif next_state in EGG_BEATER:
        if tool_collected == 0:
            return 10  
        else:
            return -10
    elif tool_collected == 0:
        if state[0] <= 4:
            distance_penalty = -manhattan_distance(next_state, EGG_BEATER[0])
        else:
            distance_penalty = -manhattan_distance(next_state, EGG_BEATER[1])
        return distance_penalty - 1 
    elif state[0] <= 4:
        distance_penalty = -manhattan_distance(next_state, GATES[0])
        return distance_penalty - 1  
    else:
        # Negative reward proportional to distance to the oven
        distance_penalty = -manhattan_distance(next_state, GOAL_STATE)
        return distance_penalty - 1 

def epsilon_greedy_policy(state, tool_collected, epsilon):
    if random.uniform(0, 1) < epsilon:
        return random.choice(ACTIONS) 
    q_values_for_state = [q_values[(state, tool_collected, action)] for action in ACTIONS]
    max_q_value = max(q_values_for_state)
    best_actions = [action for action in ACTIONS if q_values[(state, tool_collected, action)] == max_q_value]
    return random.choice(best_actions)

for episode in range(EPISODES):
    state = START_STATE
    tool_collected = 0 
    while not (state == GOAL_STATE and tool_collected == 1):
        action = epsilon_greedy_policy(state, tool_collected, EPSILON)
        next_state = get_next_state(state, action)
        reward = get_reward(state, next_state, tool_collected, action)
        next_tool_collected = tool_collected
        if next_state in EGG_BEATER:
            next_tool_collected = 1  # Collect the tool
        
        max_q_next = max(q_values[(next_state, next_tool_collected, a)] for a in ACTIONS)
        q_values[(state, tool_collected, action)] += ALPHA * (
            reward + GAMMA * max_q_next - q_values[(state, tool_collected, action)]
        )
        state, tool_collected = next_state, next_tool_collected

policy = {}
for x in range(1, GRID_SIZE[0] + 1):
    for y in range(1, GRID_SIZE[1] + 1):
        for tool_collected in [0, 1]:
            state = (x, y)
            if is_valid(state):
                q_values_for_state = [q_values[(state, tool_collected, action)] for action in ACTIONS]
                print(x, y, q_values_for_state)
                best_action = ACTIONS[np.argmax(q_values_for_state)]
                print(best_action)
                policy[(state, tool_collected)] = best_action

def draw_policy(policy, tool_collected=1):
    plt.figure(figsize=(10, 6))
    plt.xticks(range(0, 10, 1))
    plt.yticks(range(0, 5, 1))
    for ((x, y), t), action in policy.items():
        if t != tool_collected:
            continue
        # print(((x, y), t), action)
        if action == 'up':
            plt.arrow(x + 0.5, y + 0.25, 0, 0.5, head_width=0.1)
        elif action == 'down':
            plt.arrow(x + 0.5, y + 0.75, 0, -0.5, head_width=0.1)
        elif action == 'left':
            plt.arrow(x + 0.75, y + 0.5, -0.5, 0, head_width=0.1)
        elif action == 'right':
            plt.arrow(x + 0.25, y + 0.5, 0.5, 0, head_width=0.1)
        elif action == 'use_gate':
            plt.text(x + 0.5, y + 0.5, 'G', fontsize=12, ha='center', va='center', color='red')
    plt.xlim(1, GRID_SIZE[0] + 1)
    plt.ylim(1, GRID_SIZE[1] + 1)
    plt.grid()
    plt.show()

def export_transition_matrix(filename="transition_matrix.xlsx"):
    workbook = Workbook()
    
    state_tool_pairs = [
        ((x, y), tool_collected)
        for x in range(1, GRID_SIZE[0] + 1)
        for y in range(1, GRID_SIZE[1] + 1)
        for tool_collected in [0, 1]
        if is_valid((x, y))
    ]
    
    state_tool_to_index = {pair: idx for idx, pair in enumerate(state_tool_pairs)}
    
    for action in ACTIONS:
        sheet = workbook.create_sheet(title=action)
        
        # Initialize the matrix for this action
        matrix_size = len(state_tool_pairs)
        transition_matrix = [[0.0] * matrix_size for _ in range(matrix_size)]
        
        # Populate the matrix for this action
        for current_index, (current_state, tool_collected) in enumerate(state_tool_pairs):
            next_state = get_next_state(current_state, action)
            next_tool_collected = tool_collected
            if next_state in EGG_BEATER:
                next_tool_collected = 1  # Collect the tool if in egg beater
            
            if (next_state, next_tool_collected) in state_tool_to_index:
                next_index = state_tool_to_index[(next_state, next_tool_collected)]
                transition_matrix[current_index][next_index] = 1.0  # Deterministic transition
        
        header = [f"({s[0]}, {s[1]}, {tc})" for s, tc in state_tool_pairs]
        sheet.append([""] + header)
        
        for i, (current_state, tool_collected) in enumerate(state_tool_pairs):
            row = [f"({current_state[0]}, {current_state[1]}, {tool_collected})"] + transition_matrix[i]
            sheet.append(row)
        
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=matrix_size + 1, min_col=2, max_col=matrix_size + 1), start=1):
            for cell in row:
                if cell.value == 1.0:
                    cell.fill = highlight_fill
    
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    
    workbook.save(filename)
    print(f"Transition matrix exported to {filename}")

draw_policy(policy, tool_collected=1)
